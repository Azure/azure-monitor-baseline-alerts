import os
import yaml
import json
import argparse

from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

# Parse command line arguments
def parseArguments():
  parser = argparse.ArgumentParser(description='This script will export the alerts.yaml files into a single Excel file.')
  parser.add_argument('-p', '--path', type=str, required=False, metavar='path', help='Path to services directory', default='../../services')
  parser.add_argument('-x', '--output-xls', type=str, required=False, metavar='file', help='Output path to Excel file')
  parser.add_argument('-j', '--output-json', type=str, required=False, metavar='file', help='Output path to JSON file')
  parser.add_argument('-y', '--output-yaml', type=str, required=False, metavar='file', help='Output path to YAML file')
  parser.add_argument('-t', '--template', type=str, required=False, metavar='template', help='Path to Excel template', default='alerts-template.xlsx')
  parser.add_argument('-s', '--export-hidden', action='store_true', help='Export alerts rules where visible is set to false')
  args = parser.parse_args()

  return args

# Output the alerts to a JSON file
def outputToJsonFile(data, filename):
  # Write the results to a file
  with open(filename, "w") as f:
      json.dump(data, f, indent=2)

# Output the query results to a JSON file
def outputToYamlFile(data, filename):
  # Write the results to a file
  with open(filename, "w") as f:
      yaml.dump(data, f, indent=2)

def readYamlData(dir, export_hidden):

  # Walk the directory tree and load all the alerts.yaml files
  # into a list of dictionaries using the folder path as the structure
  # for the dictionary.  This will allow us to easily export the data
  # to a CSV or XLS file.
  data = {}
  for root, dirs, files in os.walk(dir):
    for file in files:
      if file == 'alerts.yaml':
        with open(os.path.join(root, file)) as f:
          # Get current folder name and parent folder name
          # to use as keys in the dictionary
          resourceType = os.path.basename(root)
          resouceCategory = os.path.basename(os.path.dirname(root))

          if not resouceCategory in data:
            data[resouceCategory] = {}

          if not resourceType in data[resouceCategory]:
            data[resouceCategory][resourceType] = []

          alerts = yaml.safe_load(f)
          if alerts:
            for alert in alerts:
              if (not export_hidden) and ('visible' in alert) and (alert['visible'] == False):
                continue
              data[resouceCategory][resourceType].append(alert)

  return data

def findColumn(ws, colName, headerRow=1):
  # Find the column for the property
  col = 0
  for cell in ws[1]:
    if cell.value.lower() == colName.lower():
      col = cell.column
      break

  return col


def addAlertToSheet(alert, ws, headerRow=1):

  # Add general alert paramters
  for key in alert:
    col = findColumn(ws, key, headerRow)
    if col > 0:
      value = ''

      if key == 'tags' and alert[key] is not None:
        value = ', '.join(alert[key])
      elif key == 'references':
        references = alert['references']
        urls = []

        if references:
          for ref in references:
            if 'url' in ref:
              urls.append(ref['url'])
            else:
              print ('No URL in reference: ' + ref['name'])
        else:
          print ('No references in alert: ' + alert['name'])

        value = '\n'.join(urls)
      elif type(alert[key]) is str or type(alert[key]) is int or type(alert[key]) is bool:
        value = alert[key]

      if value != '':
        ws.cell(row=ws.max_row, column=col).value = value

  # Add the properties
  properties = alert['properties']
  for key in properties:

    col = findColumn(ws, key, headerRow)
    if col > 0:
      value = ''

      # Add the value to the cell
      if type(properties[key]) is str:
        value = properties[key]
      else:
        value = ws.cell(row=ws.max_row, column=col).value = json.dumps(properties[key])

      if value != '':
        ws.cell(row=ws.max_row, column=col).value = value

def exportToXls(data, templateFile, outputFile):
  wb = load_workbook(templateFile)

  wsMetric = wb['Metric Alerts']
  wsLog = wb['Log Alerts']
  wsActivity = wb["Activity Alerts"]

  # Delete any rows below the header row
  for ws in [wsMetric, wsLog, wsActivity]:
    ws.delete_rows(2, wsMetric.max_row)

  for category in data:
    for type in data[category]:
      for alert in data[category][type]:
        columnsToAdd = [
          category,
          type
          ]

        if 'type' in alert:
          match alert['type'].lower():
            case 'metric':
              wsMetric.append(columnsToAdd)
              addAlertToSheet(alert, wsMetric)
            case 'log':
              wsLog.append(columnsToAdd)
              addAlertToSheet(alert, wsLog)
            case 'activitylog':
              wsActivity.append(columnsToAdd)
              addAlertToSheet(alert, wsActivity)
            case _:
              print('Unknown alert type: ' + alert['type'])
        else:
          print(f"No alert type for alert: {category} - {type} - {alert['name']}")

  for ws in [wsMetric, wsLog, wsActivity]:

    table_name = ws.title.replace(' ', '')

    if len(ws.tables) == 0:
      ws.add_table(Table(displayName=table_name))

    table = ws.tables.get(table_name)

    table.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    table.tableStyleInfo = TableStyleInfo(name='TableStyleMedium6', showRowStripes=True)

  wb.save(outputFile)

def main():

  args = parseArguments()

  data = readYamlData(args.path, args.export_hidden)

  if args.output_xls:
    exportToXls(data, args.template, args.output_xls)

  if args.output_json:
    outputToJsonFile(data, args.output_json)

  if args.output_yaml:
    outputToYamlFile(data, args.output_yaml)

if __name__ == '__main__':
  main()
